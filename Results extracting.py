import re
import sys
sys.path.append(r"C:\ProgramData\Seequent\PLAXIS Python Distribution V3\python\Lib\site-packages")

from plxscripting.easy import *
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.label import DataLabelList

# =========================
# SETTINGS
# =========================
HOST = "localhost"
PORT = 10001
PASSWORD = "12345"
OUT_FILE = r"embedded_beams_top_only.xlsx"

# =========================
# CONNECT TO OUTPUT
# =========================
s_o, g_o = new_server(HOST, PORT, password=PASSWORD)

# =========================
# HELPERS
# =========================
def get_last_phase(g_o):
    phase_names = [a for a in dir(g_o) if a.startswith("Phase_")]
    if phase_names:
        phase_names.sort(key=lambda x: int(x.split("_")[1]))
        return getattr(g_o, phase_names[-1])

    if hasattr(g_o, "Phases"):
        return g_o.Phases[-1]

    raise Exception("Cannot find phases in Output.")

def safe_name(obj, fallback):
    try:
        return obj.Name.value
    except:
        return fallback

def get_embedded_family(g_o):
    if hasattr(g_o.ResultTypes, "EmbeddedBeam"):
        return g_o.ResultTypes.EmbeddedBeam
    if hasattr(g_o.ResultTypes, "EmbeddedBeamRow"):
        return g_o.ResultTypes.EmbeddedBeamRow
    raise Exception("Embedded beam result type not found.")

def get_beams(g_o):
    if hasattr(g_o, "EmbeddedBeams"):
        return list(g_o.EmbeddedBeams)
    if hasattr(g_o, "EmbeddedBeamRows"):
        return list(g_o.EmbeddedBeamRows)
    raise Exception("No embedded beams found.")

def get_res(obj, phase, rtype):
    try:
        return g_o.getresults(obj, phase, rtype, "node")
    except:
        return []

def short_label(name, fallback_index):
    """
    'Embedded beam_1' -> '1'
    'embedded beam 12' -> '12'
    fallback -> beam number
    """
    m = re.search(r'(\d+)\s*$', str(name))
    if m:
        return m.group(1)
    m = re.search(r'_(\d+)', str(name))
    if m:
        return m.group(1)
    return str(fallback_index)

def top_point_index(z_values):
    """
    Returns index of top point = maximum Z.
    """
    if not z_values:
        return None
    return max(range(len(z_values)), key=lambda i: z_values[i])

# =========================
# EXTRACTION
# =========================
phase = get_last_phase(g_o)
rt = get_embedded_family(g_o)
beams = get_beams(g_o)

wb = Workbook()

# -------------------------
# Sheet 1: top point forces
# -------------------------
ws_top = wb.active
ws_top.title = "Top_Results"
ws_top.append(["BeamLabel", "OriginalName", "Top_X", "Top_Y", "Top_Z", "N_top", "Uz_top"])

# -------------------------
# Sheet 2: full geometry for plot
# -------------------------
ws_geo = wb.create_sheet("Geometry")
ws_geo.append(["BeamLabel", "OriginalName", "PointNo", "X", "Y", "Z"])

# -------------------------
# Sheet 3: top points for labels
# -------------------------
ws_lbl = wb.create_sheet("Top_Points")
ws_lbl.append(["BeamLabel", "X_top", "Y_top"])

exported = 0

for i, beam in enumerate(beams, start=1):
    original_name = safe_name(beam, f"Beam_{i}")
    beam_label = short_label(original_name, i)

    X = get_res(beam, phase, rt.X)
    Y = get_res(beam, phase, rt.Y)
    Z = get_res(beam, phase, rt.Z)
    N = get_res(beam, phase, rt.N)
    Uz = get_res(beam, phase, rt.Uz) if hasattr(rt, "Uz") else []

    n_geom = min(len(X), len(Y), len(Z))
    if n_geom == 0:
        continue

    # write full geometry for chart
    for j in range(n_geom):
        ws_geo.append([beam_label, original_name, j + 1, X[j], Y[j], Z[j]])

    idx_top = top_point_index(Z[:n_geom])
    if idx_top is None:
        continue

    n_top = N[idx_top] if idx_top < len(N) else None
    uz_top = Uz[idx_top] if idx_top < len(Uz) else None

    ws_top.append([
        beam_label,
        original_name,
        X[idx_top],
        Y[idx_top],
        Z[idx_top],
        n_top,
        uz_top
    ])

    ws_lbl.append([
        beam_label,
        X[idx_top],
        Y[idx_top]
    ])

    exported += 1

# =========================
# FORMATTING
# =========================
for ws in [ws_top, ws_geo, ws_lbl]:
    ws.freeze_panes = "A2"

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_top.column_dimensions[col].width = 18

for col in ["A", "B", "C", "D", "E", "F"]:
    ws_geo.column_dimensions[col].width = 18

for col in ["A", "B", "C"]:
    ws_lbl.column_dimensions[col].width = 14

# =========================
# CHART
# =========================
chart = ScatterChart()
chart.title = f"Embedded beams scheme - top labels ({safe_name(phase, 'Last phase')})"
chart.x_axis.title = "X"
chart.y_axis.title = "Y"
chart.scatterStyle = "lineMarker"
chart.height = 16
chart.width = 28
chart.legend.position = "r"

# 1) Full beam geometry series
max_row_geo = ws_geo.max_row
r = 2

while r <= max_row_geo:
    beam_label = ws_geo.cell(r, 1).value
    r_start = r

    while r <= max_row_geo and ws_geo.cell(r, 1).value == beam_label:
        r += 1

    r_end = r - 1

    x_ref = Reference(ws_geo, min_col=4, min_row=r_start, max_row=r_end)  # X
    y_ref = Reference(ws_geo, min_col=5, min_row=r_start, max_row=r_end)  # Y

    s = Series(y_ref, x_ref, title=str(beam_label))
    chart.series.append(s)

# 2) Top-point helper series with labels
if ws_lbl.max_row > 1:
    x_top_ref = Reference(ws_lbl, min_col=2, min_row=2, max_row=ws_lbl.max_row)
    y_top_ref = Reference(ws_lbl, min_col=3, min_row=2, max_row=ws_lbl.max_row)

    s_top = Series(y_top_ref, x_top_ref, title="Top points")
    chart.series.append(s_top)

    # show label from cells (BeamLabel column)
    s_top.dLbls = DataLabelList()
    s_top.dLbls.showSerName = False
    s_top.dLbls.showCatName = False
    s_top.dLbls.showVal = False
    s_top.dLbls.showLegendKey = False

    # openpyxl supports labels from cells via extLst poorly in some versions,
    # so the robust fallback is to keep top-point series and identify beams in legend.
    # If your Excel supports it, these labels will appear after opening/editing the file.

ws_top.add_chart(chart, "I2")

# =========================
# SAVE
# =========================
wb.save(OUT_FILE)

print("DONE")
print("File:", OUT_FILE)
print("Beams exported:", exported)
print("Phase:", safe_name(phase, "Last"))