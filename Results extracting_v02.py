import re
import math
import sys
sys.path.append(r"C:\ProgramData\Seequent\PLAXIS Python Distribution V3\python\Lib\site-packages")

from plxscripting.easy import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from math import ceil

# =========================
# SETTINGS
# =========================
HOST = "localhost"
PORT = 10001
PASSWORD = "12345"

EXCEL_FILE = r"embedded_beams_top_results.xlsx"
PLOT_FILE_PNG = r"embedded_beams_scheme.png"

# softest -> stiffest
COLOR_PALETTE = [
    ("1F4E79", "#1F4E79"),  # 1 softest - dark blue
    ("2F75B5", "#2F75B5"),  # 2 blue
    ("5B9BD5", "#5B9BD5"),  # 3 light blue
    ("00B0F0", "#00B0F0"),  # 4 cyan
    ("70AD47", "#70AD47"),  # 5 green
    ("A9D18E", "#A9D18E"),  # 6 light green
    ("FFD966", "#FFD966"),  # 7 yellow
    ("F4B183", "#F4B183"),  # 8 light orange
    ("ED7D31", "#ED7D31"),  # 9 orange
    ("C00000", "#C00000"),  # 10 stiffest - red
]

# =========================
# CONNECT TO OUTPUT
# =========================
s_o, g_o = new_server(HOST, PORT, password=PASSWORD)

# =========================
# HELPERS
# =========================
def get_last_phase(g_o):
    phase_names = [a for a in dir(g_o) if a.startswith("Phase_")]
    if phase_names:
        phase_names.sort(key=lambda x: int(x.split("_")[1]))
        return getattr(g_o, phase_names[-1])

    if hasattr(g_o, "Phases"):
        return g_o.Phases[-1]

    raise Exception("Cannot find phases in Output.")

def safe_name(obj, fallback):
    try:
        return obj.Name.value
    except:
        return fallback

def get_embedded_family(g_o):
    if hasattr(g_o.ResultTypes, "EmbeddedBeam"):
        return g_o.ResultTypes.EmbeddedBeam
    if hasattr(g_o.ResultTypes, "EmbeddedBeamRow"):
        return g_o.ResultTypes.EmbeddedBeamRow
    raise Exception("Embedded beam result type not found.")

def get_beams(g_o):
    if hasattr(g_o, "EmbeddedBeams"):
        return list(g_o.EmbeddedBeams)
    if hasattr(g_o, "EmbeddedBeamRows"):
        return list(g_o.EmbeddedBeamRows)
    raise Exception("No embedded beams found.")

def get_res(obj, phase, rtype):
    try:
        return g_o.getresults(obj, phase, rtype, "node")
    except:
        return []

def short_label(name, fallback_index):
    s = str(name)
    m = re.search(r'_(\d+)$', s)
    if m:
        return m.group(1)
    m = re.search(r'(\d+)$', s)
    if m:
        return m.group(1)
    return str(fallback_index)

def top_point_index(z_values):
    if not z_values:
        return None
    return max(range(len(z_values)), key=lambda i: z_values[i])

def beam_length_3d(x_vals, y_vals, z_vals):
    total = 0.0
    for i in range(1, len(x_vals)):
        dx = x_vals[i] - x_vals[i - 1]
        dy = y_vals[i] - y_vals[i - 1]
        dz = z_vals[i] - z_vals[i - 1]
        total += math.sqrt(dx * dx + dy * dy + dz * dz)
    return total

def calc_kz(n_top, uz_top):
    if n_top is None or uz_top is None or uz_top == 0:
        return None
    return n_top / uz_top

def assign_kz_classes(beam_plot_data):
    """
    Base rule: ceil(number_of_beams / 10)
    Improvement: if there are at least 2 beams, use at least 2 classes/colors
    Lowest Kz = softest
    Highest Kz = stiffest
    """
    n_beams = len(beam_plot_data)
    if n_beams == 0:
        return 0

    n_classes = ceil(n_beams / 10)

    if n_beams >= 2:
        n_classes = max(2, n_classes)
    else:
        n_classes = 1

    n_classes = min(n_classes, len(COLOR_PALETTE))

    valid = [b for b in beam_plot_data if b["Kz"] is not None]

    for b in beam_plot_data:
        b["kz_class"] = None
        b["class_name"] = "No Kz"
        b["excel_color"] = "D9D9D9"
        b["plot_color"] = "#D9D9D9"

    if not valid:
        return n_classes

    valid_sorted = sorted(valid, key=lambda b: b["Kz"])
    total_valid = len(valid_sorted)

    for idx, beam in enumerate(valid_sorted):
        class_id = int(idx * n_classes / total_valid)
        if class_id >= n_classes:
            class_id = n_classes - 1

        excel_color, plot_color = COLOR_PALETTE[class_id]
        beam["kz_class"] = class_id + 1
        beam["class_name"] = f"Class {class_id + 1}"
        beam["excel_color"] = excel_color
        beam["plot_color"] = plot_color

    return n_classes

def beam_sort_key(b):
    try:
        return (0, int(b["beam_id"]))
    except:
        return (1, str(b["beam_id"]))

# =========================
# READ DATA
# =========================
phase = get_last_phase(g_o)
rt = get_embedded_family(g_o)
beams = get_beams(g_o)

beam_plot_data = []

for i, beam in enumerate(beams, start=1):
    original_name = safe_name(beam, f"Embedded beam_{i}")
    beam_id = short_label(original_name, i)

    X = get_res(beam, phase, rt.X)
    Y = get_res(beam, phase, rt.Y)
    Z = get_res(beam, phase, rt.Z)
    N = get_res(beam, phase, rt.N)
    Uz = get_res(beam, phase, rt.Uz) if hasattr(rt, "Uz") else []

    n_geom = min(len(X), len(Y), len(Z))
    if n_geom == 0:
        continue

    X = X[:n_geom]
    Y = Y[:n_geom]
    Z = Z[:n_geom]

    idx_top = top_point_index(Z)
    if idx_top is None:
        continue

    n_top = N[idx_top] if idx_top < len(N) else None
    uz_top = Uz[idx_top] if idx_top < len(Uz) else None
    length_3d = beam_length_3d(X, Y, Z)
    kz = calc_kz(n_top, uz_top)

    beam_plot_data.append({
        "beam_id": beam_id,
        "original_name": original_name,
        "X": X,
        "Y": Y,
        "Z": Z,
        "top_x": X[idx_top],
        "top_y": Y[idx_top],
        "top_z": Z[idx_top],
        "N_top": n_top,
        "Uz_top": uz_top,
        "Length": length_3d,
        "Kz": kz,
    })

# assign classes/colors
n_classes = assign_kz_classes(beam_plot_data)

# =========================
# CLASS AVERAGES
# =========================
class_kz_values = {i: [] for i in range(1, n_classes + 1)}

for beam in beam_plot_data:
    if beam["Kz"] is not None and beam["kz_class"] is not None:
        class_kz_values[beam["kz_class"]].append(beam["Kz"])

class_kz_avg = {}
for cls, values in class_kz_values.items():
    if values:
        class_kz_avg[cls] = sum(values) / len(values)
    else:
        class_kz_avg[cls] = None

# sort output rows
beam_plot_data.sort(key=beam_sort_key)

# =========================
# EXPORT TO EXCEL
# =========================
wb = Workbook()

ws1 = wb.active
ws1.title = "Top_Results"
ws1.append([
    "BeamID",
    "OriginalName",
    "Top_X",
    "Top_Y",
    "Top_Z",
    "N_top",
    "Uz_top",
    "Length",
    "Kz (kN/m)",
    "Stiffness Class",
])

for beam in beam_plot_data:
    ws1.append([
        beam["beam_id"],
        beam["original_name"],
        beam["top_x"],
        beam["top_y"],
        beam["top_z"],
        beam["N_top"],
        beam["Uz_top"],
        beam["Length"],
        beam["Kz"],
        beam["class_name"],
    ])

for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
    ws1.column_dimensions[col].width = 18

ws1.freeze_panes = "A2"

dark_colors = {"1F4E79", "2F75B5", "C00000"}

for excel_row, beam in enumerate(beam_plot_data, start=2):
    fill = PatternFill(
        fill_type="solid",
        start_color=beam["excel_color"],
        end_color=beam["excel_color"]
    )
    ws1[f"I{excel_row}"].fill = fill
    ws1[f"J{excel_row}"].fill = fill

    if beam["excel_color"] in dark_colors:
        ws1[f"I{excel_row}"].font = Font(color="FFFFFF", bold=True)
        ws1[f"J{excel_row}"].font = Font(color="FFFFFF", bold=True)

ws2 = wb.create_sheet("Geometry")
ws2.append(["BeamID", "OriginalName", "PointNo", "X", "Y", "Z"])

for beam in beam_plot_data:
    for j, (x, y, z) in enumerate(zip(beam["X"], beam["Y"], beam["Z"]), start=1):
        ws2.append([beam["beam_id"], beam["original_name"], j, x, y, z])

for col in ["A", "B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 18

ws2.freeze_panes = "A2"

ws3 = wb.create_sheet("Legend")
ws3.append(["Class", "Meaning", "Color"])

for i in range(n_classes):
    excel_color, _ = COLOR_PALETTE[i]
    class_name = f"Class {i + 1}"

    if n_classes == 1:
        meaning = "All beams"
    elif i == 0:
        meaning = "Softest"
    elif i == n_classes - 1:
        meaning = "Stiffest"
    else:
        meaning = "Intermediate"

    ws3.append([class_name, meaning, ""])
    fill = PatternFill(fill_type="solid", start_color=excel_color, end_color=excel_color)
    ws3[f"C{i + 2}"].fill = fill

    if excel_color in dark_colors:
        ws3[f"A{i + 2}"].font = Font(color="FFFFFF", bold=True)
        ws3[f"B{i + 2}"].font = Font(color="FFFFFF", bold=True)
        ws3[f"C{i + 2}"].font = Font(color="FFFFFF", bold=True)
        ws3[f"A{i + 2}"].fill = fill
        ws3[f"B{i + 2}"].fill = fill

for col in ["A", "B", "C"]:
    ws3.column_dimensions[col].width = 18

ws4 = wb.create_sheet("Class_Averages")
ws4.append(["Class", "Average Kz (kN/m)", "Number of Beams"])

for cls in range(1, n_classes + 1):
    avg_kz = class_kz_avg.get(cls)
    count = len(class_kz_values.get(cls, []))

    ws4.append([f"Class {cls}", avg_kz, count])

    excel_color, _ = COLOR_PALETTE[cls - 1]
    fill = PatternFill(fill_type="solid", start_color=excel_color, end_color=excel_color)

    row = cls + 1
    ws4[f"A{row}"].fill = fill
    ws4[f"B{row}"].fill = fill
    ws4[f"C{row}"].fill = fill

    if excel_color in dark_colors:
        ws4[f"A{row}"].font = Font(color="FFFFFF", bold=True)
        ws4[f"B{row}"].font = Font(color="FFFFFF", bold=True)
        ws4[f"C{row}"].font = Font(color="FFFFFF", bold=True)

for col in ["A", "B", "C"]:
    ws4.column_dimensions[col].width = 22

wb.save(EXCEL_FILE)

# =========================
# MATPLOTLIB GRAPHICAL OUTPUT
# =========================
plt.figure(figsize=(14, 10))

for beam in beam_plot_data:
    x = beam["X"]
    y = beam["Y"]
    top_x = beam["top_x"]
    top_y = beam["top_y"]
    plot_color = beam["plot_color"]

    plt.plot(x, y, linewidth=2.2, color=plot_color)
    plt.scatter([top_x], [top_y], s=55, color=plot_color, edgecolors="black", linewidths=0.7)

    label = f"{beam['beam_id']} (C{beam['kz_class']})" if beam["kz_class"] else beam["beam_id"]

    plt.annotate(
        label,
        (top_x, top_y),
        xytext=(6, 6),
        textcoords="offset points",
        fontsize=8.5,
        bbox=dict(boxstyle="round,pad=0.2", fc="white", alpha=0.7, lw=0)
    )

# custom legend with average Kz
legend_elements = []
for cls in range(1, n_classes + 1):
    _, plot_color = COLOR_PALETTE[cls - 1]
    avg_kz = class_kz_avg.get(cls)
    count = len(class_kz_values.get(cls, []))

    if avg_kz is None:
        legend_label = f"C{cls}: no data"
    else:
        legend_label = f"C{cls}: avg Kz = {avg_kz:.2f} kN/m (n={count})"

    legend_elements.append(
        Line2D([0], [0], color=plot_color, lw=3, label=legend_label)
    )

plt.legend(
    handles=legend_elements,
    title="Stiffness classes",
    loc="best",
    fontsize=8,
    title_fontsize=9,
    framealpha=0.9
)

plt.xlabel("X")
plt.ylabel("Y")
plt.title(f"Embedded beams stiffness classes - top points labeled ({safe_name(phase, 'Last phase')})")
plt.grid(True, alpha=0.3)
plt.axis("equal")
plt.tight_layout()
plt.savefig(PLOT_FILE_PNG, dpi=300, bbox_inches="tight")
plt.show()

print("DONE")
print("Excel file:", EXCEL_FILE)
print("Plot file:", PLOT_FILE_PNG)
print("Beams exported:", len(beam_plot_data))
print("Number of stiffness colors/classes:", n_classes)
print("Phase:", safe_name(phase, "Last"))