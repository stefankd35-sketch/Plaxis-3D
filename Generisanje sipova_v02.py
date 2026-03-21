import sys
sys.path.append(r"C:\ProgramData\Seequent\PLAXIS Python Distribution V3\python\Lib\site-packages")

from pathlib import Path
from openpyxl import load_workbook
from plxscripting.easy import *

# ============================================================
# USER SETTINGS
# ============================================================
PLAXIS_HOST = "localhost"
PLAXIS_PORT = 10000
PLAXIS_PASSWORD = "12345"

EXCEL_FILE = Path(r"C:\Users\jelen\OneDrive\Desktop\pyPlaxis3D\pile_grid.xlsx")
SHEET_NAME = "GridInput"

MAX_ROWS = 30
MAX_COLS = 30
RENAME_OBJECTS = True

# ============================================================
# CONNECT TO PLAXIS
# ============================================================
s_i, g_i = new_server(PLAXIS_HOST, PLAXIS_PORT, password=PLAXIS_PASSWORD)

# ============================================================
# HELPERS
# ============================================================
def to_float(value, name):
    try: return float(value)
    except: raise ValueError(f"Invalid numeric: {name}")

def to_int(value, name):
    try: return int(float(value))
    except: raise ValueError(f"Invalid integer: {name}")

def read_grid_parameters(excel_path, sheet_name):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    params = {str(row[0]).strip(): row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    wb.close()
    
    return {
        "start_x": to_float(params["start_x"], "start_x"),
        "start_y": to_float(params["start_y"], "start_y"),
        "spacing_x": to_float(params["spacing_x"], "spacing_x"),
        "spacing_y": to_float(params["spacing_y"], "spacing_y"),
        "rows": to_int(params["rows"], "rows"),
        "cols": to_int(params["cols"], "cols"),
        "z_top": to_float(params["z_top"], "z_top"),
        "z_bottom": to_float(params["z_bottom"], "z_bottom"),
        "name_prefix": str(params.get("name_prefix", "P")).strip(),
        "material_name": str(params["material_name"]).strip(),
    }

def find_embedded_beam_material_by_name(g_i, material_name):
    name_lower = material_name.lower().strip()
    # Looking in the global Materials list is the most robust way
    for mat in g_i.Materials[:]:
        try:
            if mat.Name.value.lower().strip() == name_lower:
                return mat
        except: continue
    raise ValueError(f"Material '{material_name}' not found in PLAXIS.")

def create_embedded_pile(g_i, name, x, y, z_top, z_bottom):
    # This returns: [Point_17, Point_18, Line_1, EmbeddedBeam_9]
    results = g_i.embeddedbeam((x, y, z_top), (x, y, z_bottom))
    
    actual_pile_link = None
    
    # Loop through all objects created
    for obj in results:
        try:
            # CRITICAL: Target the 'EmbeddedBeam' type, not the 'Line'
            if obj.TypeName.value == "EmbeddedBeam":
                actual_pile_link = obj
                break
        except:
            continue
            
    # Fallback if the loop fails
    if not actual_pile_link:
        actual_pile_link = results[-1] # Usually the last item in the list

    if RENAME_OBJECTS:
        try:
            g_i.rename(actual_pile_link, name)
        except:
            pass

    return actual_pile_link

def assign_material_to_pile(g_i, pile_line, material):
    try:
        # This will now result in: setmaterial Line_XX Pile_Material_01
        g_i.setmaterial(pile_line, material)
        print(f"   -> Assigned {material.Name.value} to {pile_line.Name.value}")
    except Exception as e:
        print(f"   -> Assignment failed: {e}")
    


# ============================================================
# MAIN EXECUTION
# ============================================================
grid = read_grid_parameters(EXCEL_FILE, SHEET_NAME)
material = find_embedded_beam_material_by_name(g_i, grid["material_name"])

print(f"Ready to create {grid['rows'] * grid['cols']} piles with material: {grid['material_name']}")
g_i.gotostructures()

created_count = 0
errors = []

for r in range(1, grid["rows"] + 1):
    for c in range(1, grid["cols"] + 1):
        name = f"{grid['name_prefix']}_{r:02d}_{c:02d}"
        x = grid["start_x"] + (c - 1) * grid["spacing_x"]
        y = grid["start_y"] + (r - 1) * grid["spacing_y"]
        
        try:
            # 1. Create
            new_pile = create_embedded_pile(g_i, name, x, y, grid["z_top"], grid["z_bottom"])
            
            # 2. Assign Material (Passing g_i, pile, and material)
            assign_material_to_pile(g_i, new_pile, material)
            
            created_count += 1
            print(f"Success: {name} at ({x}, {y})")
        except Exception as e:
            errors.append(f"{name}: {str(e)}")
            print(f"Error on {name}: {e}")

print("-" * 30)
print(f"Finished. Created {created_count} piles.")
if errors:
    print("Errors encountered:")
    for err in errors: print(f" - {err}")

# ============================================================
# HELPERS
# ============================================================
def to_float(value, name):
    try: return float(value)
    except: raise ValueError(f"Invalid numeric value for '{name}': {value}")

def read_slab_parameters(excel_path, sheet_name):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    # Creates a dictionary from Column A and Column B
    params = {str(row[0]).strip(): row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    wb.close()
    
    return {
        "x0": to_float(params["slab_x"], "slab_x"),
        "y0": to_float(params["slab_y"], "slab_y"),
        "w": to_float(params["slab_width"], "slab_width"),
        "l": to_float(params["slab_length"], "slab_length"),
        "z": to_float(params["slab_z"], "slab_z"),
        "mat_name": str(params["plate_material"]).strip(),
    }

def find_plate_material(g_i, material_name):
    for mat in g_i.Materials[:]:
        # Filter for Plate Materials specifically
        if mat.TypeName.value == "PlateMat" and mat.Name.value.lower() == material_name.lower():
            return mat
    raise ValueError(f"Plate Material '{material_name}' not found in the project.")

# ============================================================
# PLATE DEFINITION
# ============================================================
try:
    # 1. Connect and Setup
    s_i, g_i = new_server(PLAXIS_HOST, PLAXIS_PORT, password=PLAXIS_PASSWORD)
    g_i.gotostructures()

    # 2. Load Data
    data = read_slab_parameters(EXCEL_FILE, SHEET_NAME)
    
    # 3. Geometry Logic
    # We define the 4 corners of the rectangle in order
    p1 = (data["x0"], data["y0"], data["z"])
    p2 = (data["x0"] + data["w"], data["y0"], data["z"])
    p3 = (data["x0"] + data["w"], data["y0"] + data["l"], data["z"])
    p4 = (data["x0"], data["y0"] + data["l"], data["z"])

    print(f"Drawing slab at Z = {data['z']}...")

    # 4. Create Objects
    # Create the surface using the 4 calculated points
    new_surface = g_i.surface(p1, p2, p3, p4)
    
    # Turn that surface into a Plate object
    new_plate = g_i.plate(new_surface)
    
    # 5. Material Assignment
    mat_obj = find_plate_material(g_i, data["mat_name"])
    g_i.setmaterial(new_plate, mat_obj)

    print("-----------------------------------------------")
    print(f"SUCCESS: Slab created!")
    print(f"Dimensions: {data['w']}m x {data['l']}m")
    print(f"Location: Start({data['x0']}, {data['y0']}) to End({data['x0']+data['w']}, {data['y0']+data['l']})")
    print(f"Material: {mat_obj.Name.value}")
    print("-----------------------------------------------")

except Exception as e:
    print(f"CRITICAL ERROR: {e}")

# 2. CONDITIONAL: SURFACE AT Z=0
if data["z"] < 0:
    print(f"Elevation {data['z']} is below 0. Creating influence surface at Z=0...")
    
    # Calculate expansion. 
    # Using abs() to ensure '4 * elevation' results in an expansion even if Z is negative.
    expansion = 4 * abs(data["z"])
    
    new_w = data["w"] + expansion
    new_l = data["l"] + expansion
    
    # To keep the new surface centered over the slab:
    # Shift the start point back by half of the total expansion (2 * abs(z))
    new_x0 = data["x0"] - (expansion / 2)
    new_y0 = data["y0"] - (expansion / 2)
    
    # Points for the Z=0 surface
    s1 = (new_x0, new_y0, 0)
    s2 = (new_x0 + new_w, new_y0, 0)
    s3 = (new_x0 + new_w, new_y0 + new_l, 0)
    s4 = (new_x0, new_y0 + new_l, 0)
    
    # Create the top surface
    top_surface = g_i.surface(s1, s2, s3, s4)
    print(f"Success: Created 0-elevation surface with dimensions {new_w}x{new_l}")

else:
    print("Elevation is 0 or higher. No extra surface created.")

print("Finished.")